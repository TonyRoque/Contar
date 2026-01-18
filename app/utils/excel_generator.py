import pandas as pd
import logging
from typing import List, Dict, Any
from datetime import datetime
from pathlib import Path

from app.utils.constants import STATUS_PRIORITY

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """
    Gera relatórios em Excel com formatação padronizada.
    
    Recursos:
    - Ordenação semântica por status
    - Formatação profissional de células
    - Tratamento de tipos de dados
    - Logging de operações
    """

    @staticmethod
    def gerar_relatorio(
        resultados: List[Dict[str, Any]],
        caminho_saida: str,
        incluir_timestamp: bool = True
    ) -> str:
        """
        Converte lista de resultados em arquivo Excel formatado.
        
        Args:
            resultados: Lista de dicionários com resultados
            caminho_saida: Caminho onde salvar o arquivo
            incluir_timestamp: Se True, adiciona timestamp ao nome
            
        Returns:
            Caminho absoluto do arquivo gerado
            
        Raises:
            PermissionError: Se não tem permissão de escrita
            RuntimeError: Para outros erros
        """
        try:
            # Converte resultados em DataFrame
            dados = []
            for r in resultados:
                dados.append({
                    "IP": r.get("ip", "N/A"),
                    "Torre": r.get("torre", "N/A"),
                    "Status": r.get("status", "Desconhecido"),
                    "Clientes": int(r.get("clientes", 0)),
                    "Observação": r.get("erro", r.get("observacao", "")),
                    "Data/Hora": r.get("hora", datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
                })

            if not dados:
                logger.warning("Nenhum dado para exportar")
                return ""

            df = pd.DataFrame(dados)

            # Validação e limpeza de dados
            df["Clientes"] = pd.to_numeric(df["Clientes"], errors="coerce").fillna(0).astype(int)
            
            # Ordenação semântica por status (usando mapa de prioridade)
            df["_status_order"] = df["Status"].map(STATUS_PRIORITY).fillna(999)
            df = df.sort_values(
                by=["_status_order", "Clientes"],
                ascending=[True, False]
            )
            df = df.drop(columns=["_status_order"])
            
            # Prepara caminho de saída
            caminho_path = Path(caminho_saida)
            
            if incluir_timestamp and caminho_path.stem != "relatório":
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                novo_nome = f"{caminho_path.stem}_{timestamp}{caminho_path.suffix}"
                caminho_path = caminho_path.parent / novo_nome

            # Salva com openpyxl
            try:
                df.to_excel(caminho_path, index=False, engine='openpyxl')
            except ImportError:
                logger.warning("openpyxl não disponível, usando xlsxwriter")
                df.to_excel(caminho_path, index=False, engine='xlsxwriter')

            # Aplica formatação (se possível)
            ExcelGenerator._aplicar_formatacao(caminho_path, df)

            logger.info(f"✅ Relatório gerado com sucesso: {caminho_path}")
            return str(caminho_path)

        except PermissionError:
            logger.error("Erro de permissão: O arquivo Excel pode estar aberto.")
            raise PermissionError(
                "Não foi possível salvar o arquivo. "
                "Feche o Excel e tente novamente."
            )
        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {e}", exc_info=True)
            raise RuntimeError(f"Falha na geração do relatório: {e}") from e

    @staticmethod
    def _aplicar_formatacao(caminho_arquivo: Path, df: pd.DataFrame) -> None:
        """
        Aplica formatação ao arquivo Excel (linhas alternadas, larguras, etc).
        
        Args:
            caminho_arquivo: Caminho do arquivo Excel
            df: DataFrame com dados
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = load_workbook(caminho_arquivo)
            ws = wb.active

            # Define cores por status
            status_colors = {
                "Online": "C6EFCE",      # Verde claro
                "Offline": "FFC7CE",     # Vermelho claro
                "Erro": "FFEB9C",        # Amarelo claro
                "Timeout": "FFE699",     # Amarelo mais forte
                "Erro Auth": "FFC7CE",   # Vermelho claro
                "Erro SSH": "FFC7CE",    # Vermelho claro
            }

            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Formata cabeçalho
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Formata linhas de dados
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                status_cell = row[2]  # Coluna de Status
                status = status_cell.value

                # Aplica cor de fundo baseada no status
                cor = status_colors.get(status, "FFFFFF")
                fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

                for cell in row:
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Ajusta largura das colunas
            widths = {"A": 18, "B": 15, "C": 15, "D": 12, "E": 30, "F": 20}
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            wb.save(caminho_arquivo)
            logger.debug(f"✅ Formatação aplicada ao Excel")

        except ImportError:
            logger.debug("openpyxl não disponível, pulando formatação")
        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação: {e}")
            # Continua mesmo se formatação falhar